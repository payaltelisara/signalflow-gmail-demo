import hashlib
import io
import json
import secrets
import time
from datetime import timedelta
from contextlib import asynccontextmanager

import httpx
from fastapi import Cookie, Depends, FastAPI, File, Header, HTTPException, Request, Response, UploadFile, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import RedirectResponse, StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import text
from sqlalchemy.orm import Session
from redis import Redis

from .auth import CSRF_COOKIE, SESSION_COOKIE, current_user, hash_password, issue_session, require_roles, verify_csrf, verify_password, workspace_membership
from .config import get_settings
from .db import SessionLocal, engine, get_db

from .campaigns import audience_leads, campaign_data
from .gmail import GmailConfigurationError, code_verifier, encrypt_refresh_token, exchange_code, google_email, oauth_url
from .integrations import ApolloConfigurationError, apollo_people_csv, apollo_people_search, public_connection, registry
from .smtp_imap import SmtpImapConfigurationError, configured as smtp_imap_configured, send as smtp_imap_send, setup_required as smtp_imap_setup_required, verify as verify_smtp_imap
from .jobs import create_job, job_data, update_job
from .models import AiSuggestion, AuditLog, Campaign, CampaignAudienceMember, CampaignEnrollment, CampaignStep, Company, GmailOAuthState, Import, ImportRow, IntegrationConnection, Lead, LeadDecision, MailboxConnection, Membership, Message, OutboxEvent, OutreachDraft, ScheduledMessage, SessionToken, Suppression, User, WorkflowJob, WorkflowJobLog, Workspace, utcnow
from .services import LEAD_LIFECYCLE_TRANSITIONS, can_transition_lead_stage, derive_account_enrichment, map_headers, next_account_action, normalize_country, safe_csv, score_account
from .storage import download_url, ensure_bucket, put_bytes
from .tasks import audit, process_import

settings = get_settings()
rate_limit_redis = Redis.from_url(settings.redis_url, decode_responses=True)


def bootstrap() -> None:
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.email == settings.dev_admin_email.lower()).first()
        if not user:
            user = User(email=settings.dev_admin_email.lower(), password_hash=hash_password(settings.dev_admin_password))
            db.add(user); db.flush()
        workspace = db.query(Workspace).filter(Workspace.name == "SignalFlow demo workspace").first()
        if not workspace:
            workspace = Workspace(name="SignalFlow demo workspace"); db.add(workspace); db.flush()
        if not db.query(Membership).filter(Membership.workspace_id == workspace.id, Membership.user_id == user.id).first():
            db.add(Membership(workspace_id=workspace.id, user_id=user.id, role="admin"))
        db.commit()
    finally:
        db.close()


@asynccontextmanager
async def lifespan(_: FastAPI):
    ensure_bucket()
    bootstrap()
    yield


app = FastAPI(title="SignalFlow API", version="1.0.0", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=settings.cors_origin_list, allow_credentials=True, allow_methods=["*"], allow_headers=["*"])


@app.middleware("http")
async def correlation_id(request: Request, call_next):
    request.state.correlation_id = request.headers.get("X-Correlation-ID", secrets.token_hex(16))
    response = await call_next(request)
    response.headers["X-Correlation-ID"] = request.state.correlation_id
    return response


@app.middleware("http")
async def rate_limit(request: Request, call_next):
    if request.url.path.startswith("/api/"):
        identity = request.cookies.get(SESSION_COOKIE) or (request.client.host if request.client else "unknown")
        bucket = int(time.time() // 60)
        key = f"signalflow:rate:{hashlib.sha256(identity.encode()).hexdigest()[:24]}:{bucket}"
        try:
            count = rate_limit_redis.incr(key)
            if count == 1:
                rate_limit_redis.expire(key, 70)
            if count > settings.api_rate_limit_per_minute:
                return Response(content='{"detail":"Rate limit exceeded. Retry after the current minute."}', status_code=429, media_type="application/json", headers={"Retry-After": "60"})
        except Exception:
            # Availability wins for the local console; delivery has separate,
            # durable mailbox and domain limits enforced by the scheduler.
            pass
    return await call_next(request)


class LoginInput(BaseModel):
    email: str
    password: str


class MappingInput(BaseModel):
    mapping: dict[str, str]


class StartInput(BaseModel):
    workspace_id: str


class WorkspaceOnlyInput(BaseModel):
    workspace_id: str


class ExportInput(BaseModel):
    workspace_id: str
    kind: str = Field(pattern="^(processed|rejected|audit|outreach)$")


class ReviewInput(BaseModel):
    workspace_id: str
    next_action: str | None = None
    reason: str = Field(min_length=3, max_length=500)


class OutreachDraftReviewInput(BaseModel):
    workspace_id: str
    status: str = Field(pattern="^(approved|rejected)$")
    reason: str = Field(min_length=3, max_length=500)


class CampaignStepInput(BaseModel):
    position: int = Field(ge=1, le=8)
    delay_hours: int = Field(ge=0, le=24 * 30)
    subject: str = Field(min_length=3, max_length=320)
    body: str = Field(min_length=10, max_length=8000)
    facts_used: list[str] = Field(default_factory=list, max_length=12)


class CampaignInput(BaseModel):
    workspace_id: str
    name: str = Field(min_length=3, max_length=160)
    mailbox_id: str | None = None
    audience_filter: dict = Field(default_factory=dict)
    timezone: str = Field(default="UTC", max_length=64)
    business_hours: dict = Field(default_factory=lambda: {"start": 9, "end": 17, "weekdays": [0, 1, 2, 3, 4]})
    daily_limit: int = Field(default=20, ge=1, le=200)
    per_domain_limit: int = Field(default=1, ge=1, le=5)
    steps: list[CampaignStepInput] = Field(min_length=1, max_length=8)


class CampaignActionInput(BaseModel):
    workspace_id: str
    test_recipient: str | None = None


class CampaignSequenceUpdate(BaseModel):
    workspace_id: str
    goal: str = Field(min_length=3, max_length=800)
    offer: str = Field(min_length=3, max_length=1000)
    target_persona: str = Field(min_length=2, max_length=240)
    cta: str = Field(min_length=3, max_length=240)
    proof: str = Field(default="", max_length=1200)
    preset: str = Field(default="value_first", pattern="^(value_first|problem_proof|short_direct)$")
    steps: list[CampaignStepInput] = Field(min_length=1, max_length=8)


class CampaignSequenceGenerateInput(BaseModel):
    workspace_id: str
    goal: str = Field(min_length=3, max_length=800)
    offer: str = Field(min_length=3, max_length=1000)
    target_persona: str = Field(min_length=2, max_length=240)
    cta: str = Field(min_length=3, max_length=240)
    proof: str = Field(default="", max_length=1200)
    preset: str = Field(default="value_first", pattern="^(value_first|problem_proof|short_direct)$")


class CampaignSequenceOutput(BaseModel):
    steps: list[CampaignStepInput] = Field(min_length=2, max_length=5)


class CampaignDeliveryUpdate(BaseModel):
    workspace_id: str
    mailbox_id: str | None = None
    timezone: str = Field(default="UTC", max_length=64)
    business_hours: dict = Field(default_factory=lambda: {"start": 9, "end": 17, "weekdays": [0, 1, 2, 3, 4]})
    daily_limit: int = Field(default=20, ge=1, le=200)
    per_domain_limit: int = Field(default=1, ge=1, le=5)


class SuppressionInput(BaseModel):
    workspace_id: str
    email: str = Field(min_length=3, max_length=320)
    reason: str = Field(min_length=3, max_length=96)


class IntegrationConfigurationInput(BaseModel):
    workspace_id: str
    configuration: dict = Field(default_factory=dict)
    credentials: dict[str, str] | None = None


class IntegrationTestInput(BaseModel):
    workspace_id: str


class ApolloImportInput(BaseModel):
    workspace_id: str
    filters: dict = Field(default_factory=dict)
    page: int = Field(default=1, ge=1, le=500)
    per_page: int = Field(default=25, ge=1, le=100)


class LeadLifecycleInput(BaseModel):
    workspace_id: str
    target_stage: str = Field(pattern="^(normalized|enriching|enriched|verifying|verified|invalid|researching|qualified|disqualified|awaiting_approval|approved|queued|contacted|replied|positive|question|not_now|not_interested|unsubscribe|ooo|wrong_person|referral|unknown)$")
    reason: str = Field(min_length=3, max_length=500)


def csrf_protected(csrf_cookie: str | None = Cookie(default=None, alias=CSRF_COOKIE), x_csrf_token: str | None = Header(default=None)) -> None:
    verify_csrf(csrf_cookie, x_csrf_token)


def import_summary(record: Import) -> dict:
    return {"id": record.id, "kind": record.kind, "filename": record.filename, "status": record.status, "counters": record.counters or {}, "mapping": record.column_mapping or {}, "error_message": record.error_message, "created_at": record.created_at}


def mailbox_data(mailbox: MailboxConnection) -> dict:
    return {"id": mailbox.id, "provider": mailbox.provider, "email": mailbox.email, "status": mailbox.status, "scopes": mailbox.scopes or [], "last_history_id": mailbox.last_history_id, "last_sync_at": mailbox.last_sync_at, "last_error": mailbox.last_error, "created_at": mailbox.created_at}


def integration_connection_for(db: Session, workspace_id: str, provider: str) -> IntegrationConnection | None:
    return db.query(IntegrationConnection).filter(
        IntegrationConnection.workspace_id == workspace_id,
        IntegrationConnection.provider == provider,
    ).first()


def campaign_workspace_data(db: Session, campaign: Campaign) -> dict:
    members = db.query(CampaignAudienceMember).filter(CampaignAudienceMember.campaign_id == campaign.id).all()
    leads = {item.id: item for item in db.query(Lead).filter(Lead.id.in_([member.lead_id for member in members if member.lead_id])).all()} if members else {}
    companies = {item.id: item for item in db.query(Company).filter(Company.id.in_([member.company_id for member in members if member.company_id])).all()} if members else {}
    summary = {"total": len(members), "ready": 0, "researching": 0, "drafting": 0, "missing_email": 0, "suppressed": 0, "excluded": 0, "failed": 0}
    audience = []
    for member in members:
        lead = leads.get(member.lead_id)
        company = companies.get(member.company_id)
        readiness = member.readiness
        if lead:
            suggestion = db.query(AiSuggestion).filter(AiSuggestion.lead_id == lead.id).order_by(AiSuggestion.created_at.desc()).first()
            draft = db.query(OutreachDraft).filter(OutreachDraft.lead_id == lead.id).order_by(OutreachDraft.updated_at.desc()).first()
            if lead.qualification == "suppressed": readiness = "suppressed"
            elif not lead.email: readiness = "missing_email"
            elif suggestion and suggestion.status == "completed": readiness = "ready"
            elif suggestion and suggestion.status in {"queued", "running"}: readiness = "researching"
            elif lead.qualification == "qualified": readiness = "drafting"
            else: readiness = "excluded"
            member.readiness = readiness
        if not lead and company:
            readiness = "missing_email"; member.readiness = readiness
        summary[readiness if readiness in summary else "failed"] += 1
        audience.append({"id": member.id, "lead_id": member.lead_id, "selected": member.selected, "readiness": readiness, "warning": member.warning, "exclusion_reason": member.exclusion_reason, "name": lead.full_name if lead else company.name if company else None, "email": lead.email if lead else None, "company_id": lead.company_id if lead else member.company_id, "score": lead.score if lead else company.score if company else None})
    imports = db.query(Import).filter(Import.campaign_id == campaign.id).order_by(Import.created_at.desc()).all()
    mailbox = db.get(MailboxConnection, campaign.mailbox_id) if campaign.mailbox_id else None
    blockers = []
    if not summary["ready"]: blockers.append("No campaign-ready recipients yet")
    if not mailbox or mailbox.status != "connected": blockers.append("Connect a sending mailbox")
    if not campaign.test_sent_at: blockers.append("Send a test email")
    if not campaign.approved_at: blockers.append("Review and approve the campaign")
    preparing = db.query(WorkflowJob).filter(WorkflowJob.workspace_id == campaign.workspace_id, WorkflowJob.resource_type == "campaign", WorkflowJob.resource_id == campaign.id, WorkflowJob.job_type == "campaign_prepare").order_by(WorkflowJob.updated_at.desc()).first()
    import_pending = any(item.status in {"uploaded", "validated", "queued", "processing"} for item in imports)
    if preparing and preparing.status not in {"cancelled", "failed"}:
        if import_pending or summary["researching"] or summary["drafting"]:
            preparing.status = "running"; preparing.phase = "researching" if summary["researching"] else "preparing audience"
        else:
            preparing.status = "partially_completed" if summary["failed"] else "completed"; preparing.phase = "ready"
    results = {"sent": db.query(Message).filter(Message.campaign_id == campaign.id, Message.direction == "outbound", Message.delivery_status == "sent").count(), "replied": db.query(Message).filter(Message.campaign_id == campaign.id, Message.direction == "inbound").count(), "interested": db.query(Message).filter(Message.campaign_id == campaign.id, Message.reply_classification == "interested").count(), "bounced": db.query(Message).filter(Message.campaign_id == campaign.id, Message.reply_classification == "hard_bounce").count(), "unsubscribed": db.query(Message).filter(Message.campaign_id == campaign.id, Message.reply_classification == "unsubscribe").count(), "stopped": db.query(CampaignEnrollment).filter(CampaignEnrollment.campaign_id == campaign.id, CampaignEnrollment.status == "stopped").count()}
    return {**campaign_data(db, campaign), "audience": audience, "audience_summary": summary, "imports": [import_summary(item) for item in imports], "mailbox": mailbox_data(mailbox) if mailbox else None, "readiness": {"blockers": blockers, "can_launch": not blockers}, "results": results, "activity": [{"action": item.action, "created_at": item.created_at, "payload": item.payload} for item in db.query(AuditLog).filter(AuditLog.workspace_id == campaign.workspace_id, AuditLog.resource_id == campaign.id).order_by(AuditLog.created_at.desc()).limit(20).all()]}


def resource_jobs(db: Session, workspace_id: str, resource_id: str) -> list[dict]:
    jobs = db.query(WorkflowJob).filter(WorkflowJob.workspace_id == workspace_id, WorkflowJob.resource_id == resource_id).order_by(WorkflowJob.updated_at.desc()).all()
    return [job_data(item, db.query(WorkflowJobLog).filter(WorkflowJobLog.job_id == item.id).order_by(WorkflowJobLog.created_at.desc()).limit(20).all()) for item in jobs]


@app.get("/health/live")
def liveness():
    return {"status": "ok"}


@app.get("/health/ready")
def readiness():
    with engine.connect() as connection:
        connection.execute(text("SELECT 1"))
    if not settings.ollama_model:
        raise HTTPException(status_code=503, detail="OLLAMA_MODEL is required")
    try:
        response = httpx.get(f"{settings.ollama_base_url.rstrip('/')}/api/tags", timeout=min(settings.ollama_timeout_seconds, 5))
        response.raise_for_status()
        models = {item.get("name") for item in response.json().get("models", [])}
        if settings.ollama_model not in models:
            raise HTTPException(status_code=503, detail=f"Ollama model {settings.ollama_model} is not available")
    except httpx.HTTPError as exc:
        raise HTTPException(status_code=503, detail="Ollama is required but unavailable") from exc
    return {"status": "ready", "ollama_model": settings.ollama_model}


@app.post("/api/v1/auth/login")
def login(payload: LoginInput, response: Response, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == payload.email.lower()).first()
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid email or password")
    csrf = issue_session(response, db, user); db.commit()
    return {"user": {"id": user.id, "email": user.email}, "csrf_token": csrf}


@app.post("/api/v1/auth/logout", dependencies=[Depends(csrf_protected)])
def logout(response: Response, user: User = Depends(current_user), session_token: str | None = Cookie(default=None, alias=SESSION_COOKIE), db: Session = Depends(get_db)):
    if session_token:
        token_hash = hashlib.sha256(session_token.encode()).hexdigest()
        db.query(SessionToken).filter(SessionToken.token_hash == token_hash, SessionToken.user_id == user.id).delete()
        db.commit()
    response.delete_cookie(SESSION_COOKIE); response.delete_cookie(CSRF_COOKIE)
    return {"ok": True}


@app.get("/api/v1/auth/session")
def session_info(user: User = Depends(current_user), db: Session = Depends(get_db)):
    memberships = db.query(Membership).filter(Membership.user_id == user.id).all()
    return {"user": {"id": user.id, "email": user.email}, "workspaces": [{"id": item.workspace_id, "role": item.role, "name": db.get(Workspace, item.workspace_id).name} for item in memberships]}


@app.get("/api/v1/workspaces")
def workspaces(user: User = Depends(current_user), db: Session = Depends(get_db)):
    records = db.query(Membership).filter(Membership.user_id == user.id).all()
    return [{"id": item.workspace_id, "role": item.role, "name": db.get(Workspace, item.workspace_id).name} for item in records]


@app.post("/api/v1/imports", status_code=status.HTTP_201_CREATED, dependencies=[Depends(csrf_protected)])
async def create_import(workspace_id: str, import_kind: str = "leads", file: UploadFile = File(...), idempotency_key: str = Header(..., alias="Idempotency-Key"), user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    if import_kind not in {"leads", "accounts"}:
        raise HTTPException(status_code=422, detail="Import kind must be leads or accounts")
    filename = file.filename or "upload.csv"
    if not filename.lower().endswith((".csv", ".xlsx")):
        raise HTTPException(status_code=415, detail="Only CSV and XLSX files are supported")
    content = await file.read()
    if not content or len(content) > settings.max_upload_bytes:
        raise HTTPException(status_code=413, detail=f"File must be between 1 byte and {settings.max_upload_bytes} bytes")
    checksum = hashlib.sha256(content).hexdigest()
    request_hash = hashlib.sha256(f"{workspace_id}:{import_kind}:{checksum}:{filename}".encode()).hexdigest()
    existing = db.query(Import).filter(Import.workspace_id == workspace_id, Import.idempotency_key == idempotency_key).first()
    if existing:
        if existing.request_hash != request_hash:
            raise HTTPException(status_code=409, detail="Idempotency key was used with a different upload")
        return import_summary(existing)
    if filename.lower().endswith(".csv"):
        try:
            headers = next(__import__("csv").reader([content.decode("utf-8-sig").splitlines()[0]]))
        except Exception as exc:
            raise HTTPException(status_code=422, detail="CSV must be UTF-8 with a header row") from exc
    else:
        try:
            book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            headers = [str(value or "").strip() for value in next(book.active.iter_rows(values_only=True))]
        except Exception as exc:
            raise HTTPException(status_code=422, detail="XLSX must contain a readable header row") from exc
    if not headers or not any(headers) or len(headers) != len(set(headers)):
        raise HTTPException(status_code=422, detail="Upload must contain a non-empty header row with unique column names")
    # Content checksums remain audit evidence, but object keys must be unique:
    # a user may intentionally re-upload an unchanged file with a new import.
    object_key = f"uploads/{workspace_id}/{checksum}-{secrets.token_hex(8)}-{filename}"
    put_bytes(object_key, content, file.content_type or "application/octet-stream")
    record = Import(workspace_id=workspace_id, created_by=user.id, filename=filename, object_key=object_key, checksum=checksum, mime_type=file.content_type or "application/octet-stream", size_bytes=len(content), idempotency_key=idempotency_key, request_hash=request_hash, kind=import_kind, status="validated", column_mapping=map_headers(headers))
    db.add(record); db.flush()
    audit(db, workspace_id, "import.created", "import", record.id, {"filename": filename, "checksum": checksum, "kind": import_kind}, user.id)
    db.commit(); db.refresh(record)
    return import_summary(record)


@app.post("/api/v1/imports/{import_id}/mapping", dependencies=[Depends(csrf_protected)])
def set_mapping(import_id: str, payload: MappingInput, workspace_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    record = db.get(Import, import_id)
    if not record or record.workspace_id != workspace_id: raise HTTPException(status_code=404, detail="Import not found")
    membership = workspace_membership(workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    if record.status not in {"uploaded", "validated"}: raise HTTPException(status_code=409, detail="Mapping can no longer be changed")
    record.column_mapping = payload.mapping; db.commit(); return import_summary(record)


@app.post("/api/v1/imports/{import_id}/start", dependencies=[Depends(csrf_protected)])
def start_import(import_id: str, payload: StartInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    record = db.get(Import, import_id)
    if not record or record.workspace_id != payload.workspace_id: raise HTTPException(status_code=404, detail="Import not found")
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    if not settings.ollama_model:
        raise HTTPException(status_code=503, detail="Configure OLLAMA_MODEL before starting an import")
    if record.status not in {"uploaded", "validated", "failed"}: return import_summary(record)
    job = create_job(db, workspace_id=payload.workspace_id, created_by=user.id, job_type="import", name=f"Import {record.filename}", idempotency_key=f"import:{record.id}", resource_type="import", resource_id=record.id)
    record.status = "queued"; audit(db, payload.workspace_id, "import.queued", "import", record.id, {"job_id": job.id}, user.id); db.commit()
    process_import.delay(record.id)
    return {**import_summary(record), "job": job_data(job)}


@app.get("/api/v1/imports/{import_id}")
def get_import(import_id: str, workspace_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    record = db.get(Import, import_id)
    if not record or record.workspace_id != workspace_id: raise HTTPException(status_code=404, detail="Import not found")
    return import_summary(record)


@app.get("/api/v1/imports/{import_id}/errors")
def import_errors(import_id: str, workspace_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    record = db.get(Import, import_id)
    if not record or record.workspace_id != workspace_id: raise HTTPException(status_code=404, detail="Import not found")
    return [{"row_number": row.row_number, "errors": row.errors, "outcome": row.outcome} for row in db.query(ImportRow).filter(ImportRow.import_id == import_id, ImportRow.outcome.in_(["rejected", "needs_review"])).all()]


@app.get("/api/v1/jobs")
def list_jobs(workspace_id: str, resource_id: str | None = None, limit: int = 100, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    query = db.query(WorkflowJob).filter(WorkflowJob.workspace_id == workspace_id)
    if resource_id:
        query = query.filter(WorkflowJob.resource_id == resource_id)
    jobs = query.order_by(WorkflowJob.updated_at.desc()).limit(min(limit, 200)).all()
    return [job_data(job, db.query(WorkflowJobLog).filter(WorkflowJobLog.job_id == job.id).order_by(WorkflowJobLog.created_at.desc()).limit(20).all()) for job in jobs]


@app.get("/api/v1/jobs/{job_id}")
def get_job(job_id: str, workspace_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    job = db.get(WorkflowJob, job_id)
    if not job or job.workspace_id != workspace_id:
        raise HTTPException(status_code=404, detail="Job not found")
    return job_data(job, db.query(WorkflowJobLog).filter(WorkflowJobLog.job_id == job.id).order_by(WorkflowJobLog.created_at.desc()).limit(100).all())


@app.get("/api/v1/jobs/{job_id}/events")
def job_events(job_id: str, workspace_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    job = db.get(WorkflowJob, job_id)
    if not job or job.workspace_id != workspace_id:
        raise HTTPException(status_code=404, detail="Job not found")

    def stream():
        last_updated = None
        for _ in range(60):
            if last_updated:
                time.sleep(1)
            session = SessionLocal()
            try:
                fresh = session.get(WorkflowJob, job_id)
                if not fresh:
                    break
                marker = fresh.updated_at.isoformat()
                if marker == last_updated:
                    if fresh.status in {"completed", "partially_completed", "failed", "cancelled"}:
                        break
                    continue
                logs = session.query(WorkflowJobLog).filter(WorkflowJobLog.job_id == job_id).order_by(WorkflowJobLog.created_at.desc()).limit(20).all()
                yield f"event: job\ndata: {json.dumps(job_data(fresh, logs), default=str)}\n\n"
                last_updated = marker
                if fresh.status in {"completed", "partially_completed", "failed", "cancelled"}:
                    break
            finally:
                session.close()
            time.sleep(1)

    return StreamingResponse(stream(), media_type="text/event-stream", headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.post("/api/v1/jobs/{job_id}/cancel", dependencies=[Depends(csrf_protected)])
def cancel_job(job_id: str, payload: WorkspaceOnlyInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    job = db.get(WorkflowJob, job_id)
    if not job or job.workspace_id != payload.workspace_id:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.status in {"completed", "partially_completed", "failed", "cancelled"}:
        return job_data(job)
    job.cancellation_requested = True
    update_job(db, job, phase="cancellation requested", message="Cancellation requested by user")
    audit(db, payload.workspace_id, "job.cancellation_requested", "job", job.id, {}, user.id)
    db.commit(); return job_data(job)


def retry_workflow_job(job_id: str, payload: WorkspaceOnlyInput, user: User, db: Session):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    previous = db.get(WorkflowJob, job_id)
    if not previous or previous.workspace_id != payload.workspace_id:
        raise HTTPException(status_code=404, detail="Job not found")
    if previous.status not in {"failed", "cancelled", "partially_completed"}:
        raise HTTPException(status_code=409, detail="Only failed, cancelled, or partially completed jobs can be retried")
    attempt = previous.attempt + 1
    job = create_job(db, workspace_id=payload.workspace_id, created_by=user.id, job_type=previous.job_type, name=previous.name, idempotency_key=f"{previous.idempotency_key}:attempt:{attempt}", resource_type=previous.resource_type, resource_id=previous.resource_id, counters=previous.counters, details={**(previous.details or {}), "retry_of": previous.id})
    job.attempt = attempt
    if previous.job_type == "import" and previous.resource_id:
        record = db.get(Import, previous.resource_id)
        if not record: raise HTTPException(status_code=404, detail="Related import no longer exists")
        record.status = "queued"
    elif previous.job_type == "campaign_enrollment" and previous.resource_id:
        campaign = db.get(Campaign, previous.resource_id)
        if not campaign: raise HTTPException(status_code=404, detail="Related campaign no longer exists")
        campaign.status = "running"
        db.add(OutboxEvent(topic="campaign.enrollment_requested", payload={"campaign_id": campaign.id, "workspace_id": payload.workspace_id, "job_id": job.id}))
    elif previous.job_type == "gmail_sync" and previous.resource_id:
        db.add(OutboxEvent(topic="gmail.sync_requested", payload={"mailbox_id": previous.resource_id, "workspace_id": payload.workspace_id, "job_id": job.id}))
    elif previous.job_type == "ollama_lead_analysis" and previous.resource_id:
        db.add(OutboxEvent(topic="lead.ai_suggestion_requested", payload={"lead_id": previous.resource_id, "workspace_id": payload.workspace_id, "requested_by": user.id, "job_id": job.id}))
    elif previous.job_type == "ollama_account_research" and previous.resource_id:
        company = db.get(Company, previous.resource_id)
        if not company: raise HTTPException(status_code=404, detail="Related account no longer exists")
        company.enrichment_data = {**(company.enrichment_data or {}), "ai_research": {"status": "queued", "model": settings.ollama_model, "prompt_version": "account-research-v1"}}
        db.add(OutboxEvent(topic="company.ai_research_requested", payload={"company_id": company.id, "workspace_id": payload.workspace_id, "requested_by": user.id, "job_id": job.id}))
    else:
        raise HTTPException(status_code=409, detail="This job type does not support retry")
    audit(db, payload.workspace_id, "job.retried", "job", job.id, {"retry_of": previous.id, "attempt": attempt}, user.id)
    db.commit()
    if previous.job_type == "import" and previous.resource_id:
        process_import.delay(previous.resource_id)
    return job_data(job)


@app.post("/api/v1/jobs/{job_id}/retry", dependencies=[Depends(csrf_protected)])
def retry_job(job_id: str, payload: WorkspaceOnlyInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    return retry_workflow_job(job_id, payload, user, db)


@app.post("/api/v1/jobs/{job_id}/resume", dependencies=[Depends(csrf_protected)])
def resume_job(job_id: str, payload: WorkspaceOnlyInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    return retry_workflow_job(job_id, payload, user, db)


@app.get("/api/v1/leads")
def list_leads(workspace_id: str, qualification: str | None = None, next_action: str | None = None, source: str | None = None, limit: int = 50, offset: int = 0, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    query = db.query(Lead).filter(Lead.workspace_id == workspace_id)
    if qualification: query = query.filter(Lead.qualification == qualification)
    if next_action: query = query.filter(Lead.next_action == next_action)
    if source: query = query.filter(Lead.source == source)
    rows = query.order_by(Lead.updated_at.desc()).offset(offset).limit(min(limit, 100)).all()
    return [{"id": lead.id, "name": lead.full_name, "email": lead.email, "title": lead.title, "source": lead.source, "score": lead.score, "qualification": lead.qualification, "next_action": lead.next_action, "owner_id": lead.owner_id, "territory": lead.territory} for lead in rows]


@app.get("/api/v1/companies")
def list_companies(workspace_id: str, limit: int = 100, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    from .models import Company
    rows = db.query(Company).filter(Company.workspace_id == workspace_id).order_by(Company.name.asc()).limit(min(limit, 200)).all()
    return [{"id": item.id, "name": item.name, "domain": item.domain, "industry": item.industry, "employee_band": item.employee_band, "score": item.score, "qualification": item.qualification, "next_action": item.next_action, "owner_id": item.owner_id, "profile_data": item.profile_data, "enrichment_data": item.enrichment_data} for item in rows]


@app.post("/api/v1/companies/recalculate", dependencies=[Depends(csrf_protected)])
def recalculate_companies(payload: WorkspaceOnlyInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    from .models import Company
    companies = db.query(Company).filter(Company.workspace_id == payload.workspace_id).all()
    eligible = db.query(User).join(Membership, Membership.user_id == User.id).filter(Membership.workspace_id == payload.workspace_id).all()
    qualified = 0
    for company in companies:
        profile = company.profile_data or {}
        _, territory = normalize_country(profile.get("Company Country") or profile.get("Country"))
        account_data = {"company_domain": company.domain, "industry": company.industry, "employee_band": company.employee_band, "territory": territory}
        enrichment = derive_account_enrichment(account_data, profile)
        score, qualification, contributions = score_account(account_data, profile, enrichment)
        company.enrichment_data = enrichment
        owner = min(eligible, key=lambda candidate: db.query(Company).filter(Company.workspace_id == payload.workspace_id, Company.owner_id == candidate.id).count()) if qualification == "qualified" and eligible else None
        company.score = score; company.qualification = qualification; company.owner_id = owner.id if owner else None; company.next_action = next_account_action(qualification, bool(owner))
        qualified += int(qualification == "qualified")
        audit(db, payload.workspace_id, "company.recalculated", "company", company.id, {"total": score, "qualification": qualification, "contributions": contributions, "next_action": company.next_action}, user.id)
    db.commit()
    return {"accounts_recalculated": len(companies), "qualified_accounts": qualified}


@app.post("/api/v1/companies/ai-research", dependencies=[Depends(csrf_protected)])
def queue_account_ai_research(payload: WorkspaceOnlyInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    if not settings.ollama_model:
        raise HTTPException(status_code=503, detail="Configure OLLAMA_MODEL before running account research")
    from .models import Company
    queued = 0
    for company in db.query(Company).filter(Company.workspace_id == payload.workspace_id, Company.qualification == "qualified").all():
        enrichment = dict(company.enrichment_data or {})
        research = enrichment.get("ai_research", {})
        if research.get("status") in {"queued", "running", "completed"}:
            continue
        enrichment["ai_research"] = {"status": "queued", "model": settings.ollama_model, "prompt_version": "account-research-v1"}
        company.enrichment_data = enrichment
        db.add(OutboxEvent(topic="company.ai_research_requested", payload={"company_id": company.id, "workspace_id": payload.workspace_id, "requested_by": user.id}))
        audit(db, payload.workspace_id, "company.ai_research_queued", "company", company.id, {}, user.id)
        queued += 1
    db.commit()
    return {"queued": queued}


@app.get("/api/v1/leads/{lead_id}")
def lead_detail(lead_id: str, workspace_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    lead = db.get(Lead, lead_id)
    if not lead or lead.workspace_id != workspace_id: raise HTTPException(status_code=404, detail="Lead not found")
    return {"lead": {"id": lead.id, "name": lead.full_name, "email": lead.email, "title": lead.title, "source": lead.source, "stage": lead.stage, "score": lead.score, "qualification": lead.qualification, "next_action": lead.next_action, "owner_id": lead.owner_id, "territory": lead.territory, "raw_data": lead.raw_data}, "decisions": [{"kind": item.kind, "result": item.result, "version": item.version, "created_at": item.created_at} for item in db.query(LeadDecision).filter(LeadDecision.lead_id == lead_id).order_by(LeadDecision.created_at.desc()).all()], "ai_suggestions": [{"id": item.id, "status": item.status, "model": item.model, "output": item.output, "reviewer_status": item.reviewer_status, "error_message": item.error_message} for item in db.query(AiSuggestion).filter(AiSuggestion.lead_id == lead_id).all()], "outreach_drafts": [outreach_draft_data(item) for item in db.query(OutreachDraft).filter(OutreachDraft.lead_id == lead_id).order_by(OutreachDraft.version.desc()).all()], "events": [{"action": item.action, "payload": item.payload, "created_at": item.created_at} for item in db.query(AuditLog).filter(AuditLog.workspace_id == workspace_id, AuditLog.resource_id == lead_id).order_by(AuditLog.created_at.desc()).all()]}


def outreach_draft_data(draft: OutreachDraft) -> dict:
    return {"id": draft.id, "lead_id": draft.lead_id, "subject": draft.subject, "body": draft.body, "sequence": draft.sequence, "rationale": draft.rationale, "status": draft.status, "version": draft.version, "created_at": draft.created_at, "reviewed_at": draft.reviewed_at}


@app.get("/api/v1/outreach-drafts")
def list_outreach_drafts(workspace_id: str, status_filter: str | None = None, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    query = db.query(OutreachDraft).filter(OutreachDraft.workspace_id == workspace_id)
    if status_filter:
        query = query.filter(OutreachDraft.status == status_filter)
    drafts = query.order_by(OutreachDraft.updated_at.desc()).limit(100).all()
    leads = {lead.id: lead for lead in db.query(Lead).filter(Lead.id.in_([draft.lead_id for draft in drafts])).all()} if drafts else {}
    return [{**outreach_draft_data(draft), "lead_name": leads.get(draft.lead_id).full_name if leads.get(draft.lead_id) else None, "lead_email": leads.get(draft.lead_id).email if leads.get(draft.lead_id) else None} for draft in drafts]


@app.post("/api/v1/outreach-drafts/{draft_id}/review", dependencies=[Depends(csrf_protected)])
def review_outreach_draft(draft_id: str, payload: OutreachDraftReviewInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    draft = db.get(OutreachDraft, draft_id)
    if not draft or draft.workspace_id != payload.workspace_id:
        raise HTTPException(status_code=404, detail="Outreach draft not found")
    if draft.status != "draft":
        raise HTTPException(status_code=409, detail="Only unreviewed drafts can be reviewed")
    draft.status = payload.status; draft.reviewed_by = user.id; draft.reviewed_at = utcnow()
    audit(db, payload.workspace_id, f"outreach.draft_{payload.status}", "outreach_draft", draft.id, {"lead_id": draft.lead_id, "reason": payload.reason, "delivery": "not_sent"}, user.id)
    db.commit()
    return outreach_draft_data(draft)


@app.post("/api/v1/leads/{lead_id}/review", dependencies=[Depends(csrf_protected)])
def review_lead(lead_id: str, payload: ReviewInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    lead = db.get(Lead, lead_id)
    if not lead or lead.workspace_id != payload.workspace_id: raise HTTPException(status_code=404, detail="Lead not found")
    original = lead.next_action
    if payload.next_action: lead.next_action = payload.next_action
    audit(db, payload.workspace_id, "lead.reviewed", "lead", lead.id, {"original_next_action": original, "next_action": lead.next_action, "reason": payload.reason}, user.id)
    db.commit(); return {"id": lead.id, "next_action": lead.next_action}


@app.post("/api/v1/leads/{lead_id}/lifecycle", dependencies=[Depends(csrf_protected)])
def transition_lead_lifecycle(lead_id: str, payload: LeadLifecycleInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    lead = db.get(Lead, lead_id)
    if not lead or lead.workspace_id != payload.workspace_id:
        raise HTTPException(status_code=404, detail="Lead not found")
    current_stage = "imported" if lead.stage in {None, "", "new"} else lead.stage
    if not can_transition_lead_stage(lead.stage, payload.target_stage):
        allowed = sorted(LEAD_LIFECYCLE_TRANSITIONS.get(current_stage, set()))
        raise HTTPException(status_code=409, detail={"message": "Lifecycle transition is not allowed", "current_stage": current_stage, "allowed_next_stages": allowed})
    lead.stage = payload.target_stage
    transition = {"from": current_stage, "to": payload.target_stage, "reason": payload.reason, "source": "manual"}
    db.add(LeadDecision(lead_id=lead.id, kind="lifecycle", result=transition))
    audit(db, payload.workspace_id, "lead.lifecycle_transitioned", "lead", lead.id, transition, user.id)
    db.commit()
    return {"id": lead.id, "stage": lead.stage, "transition": transition}


@app.post("/api/v1/leads/{lead_id}/ai-retry", dependencies=[Depends(csrf_protected)])
def retry_lead_ai(lead_id: str, payload: WorkspaceOnlyInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    if not settings.ollama_model:
        raise HTTPException(status_code=503, detail="Configure OLLAMA_MODEL before retrying AI analysis")
    lead = db.get(Lead, lead_id)
    if not lead or lead.workspace_id != payload.workspace_id:
        raise HTTPException(status_code=404, detail="Lead not found")
    if lead.qualification != "qualified" or not lead.owner_id or not lead.email:
        raise HTTPException(status_code=409, detail="Only qualified, assigned contacts with an email can run AI outreach analysis")
    db.add(OutboxEvent(topic="lead.ai_suggestion_requested", payload={"lead_id": lead.id, "workspace_id": payload.workspace_id, "requested_by": user.id}))
    audit(db, payload.workspace_id, "ai.suggestion_retry_queued", "lead", lead.id, {}, user.id)
    db.commit()
    return {"lead_id": lead.id, "status": "queued"}


@app.get("/api/v1/campaigns")
def list_campaigns(workspace_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    return [campaign_data(db, campaign) for campaign in db.query(Campaign).filter(Campaign.workspace_id == workspace_id).order_by(Campaign.updated_at.desc()).all()]


@app.post("/api/v1/campaigns", status_code=status.HTTP_201_CREATED, dependencies=[Depends(csrf_protected)])
def create_campaign(payload: CampaignInput, idempotency_key: str = Header(..., alias="Idempotency-Key"), user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    fingerprint = hashlib.sha256(json.dumps(payload.model_dump(mode="json"), sort_keys=True).encode()).hexdigest()[:32]
    existing = db.query(Campaign).filter(Campaign.workspace_id == payload.workspace_id, Campaign.name == payload.name).first()
    if existing and (existing.audience_filter or {}).get("creation_idempotency_key") == idempotency_key:
        return campaign_data(db, existing)
    if existing:
        raise HTTPException(status_code=409, detail="A campaign with this name already exists")
    if payload.mailbox_id:
        mailbox = db.get(MailboxConnection, payload.mailbox_id)
        if not mailbox or mailbox.workspace_id != payload.workspace_id:
            raise HTTPException(status_code=422, detail="Selected mailbox does not belong to this workspace")
    positions = [step.position for step in payload.steps]
    if sorted(positions) != list(range(1, len(positions) + 1)):
        raise HTTPException(status_code=422, detail="Campaign steps must use consecutive positions starting at 1")
    campaign = Campaign(workspace_id=payload.workspace_id, created_by=user.id, mailbox_id=payload.mailbox_id, name=payload.name, audience_filter={**payload.audience_filter, "campaign_first": True, "creation_idempotency_key": idempotency_key, "creation_fingerprint": fingerprint}, timezone=payload.timezone, business_hours=payload.business_hours, daily_limit=payload.daily_limit, per_domain_limit=payload.per_domain_limit)
    db.add(campaign); db.flush()
    for step in payload.steps:
        db.add(CampaignStep(campaign_id=campaign.id, position=step.position, delay_hours=step.delay_hours, subject=step.subject, body=step.body, facts_used=step.facts_used))
    audit(db, payload.workspace_id, "campaign.created", "campaign", campaign.id, {"steps": len(payload.steps), "delivery": "not_started"}, user.id)
    db.commit(); db.refresh(campaign)
    return campaign_data(db, campaign)


@app.get("/api/v1/campaigns/{campaign_id}")
def get_campaign(campaign_id: str, workspace_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    campaign = db.get(Campaign, campaign_id)
    if not campaign or campaign.workspace_id != workspace_id:
        raise HTTPException(status_code=404, detail="Campaign not found")
    return {**campaign_data(db, campaign), "jobs": resource_jobs(db, workspace_id, campaign.id), "scheduled_messages": [{"id": item.id, "status": item.status, "step_position": item.step_position, "send_at": item.send_at, "error_message": item.error_message} for item in db.query(ScheduledMessage).filter(ScheduledMessage.campaign_id == campaign.id).order_by(ScheduledMessage.send_at.desc()).limit(100).all()]}


@app.get("/api/v1/campaigns/{campaign_id}/workspace")
def campaign_workspace(campaign_id: str, workspace_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    campaign = db.get(Campaign, campaign_id)
    if not campaign or campaign.workspace_id != workspace_id:
        raise HTTPException(status_code=404, detail="Campaign not found")
    result = campaign_workspace_data(db, campaign)
    db.commit()
    return result


@app.post("/api/v1/campaigns/{campaign_id}/audience/import", dependencies=[Depends(csrf_protected)])
async def import_campaign_audience(campaign_id: str, workspace_id: str, kind: str = "auto", file: UploadFile = File(...), idempotency_key: str = Header(..., alias="Idempotency-Key"), user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    campaign = db.get(Campaign, campaign_id)
    if not campaign or campaign.workspace_id != workspace_id:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if kind not in {"auto", "accounts", "leads"}:
        raise HTTPException(status_code=422, detail="Import kind must be auto, accounts, or leads")
    content = await file.read(); await file.seek(0)
    filename = file.filename or "upload.csv"
    try:
        if filename.lower().endswith(".csv"):
            headers = next(__import__("csv").reader([content.decode("utf-8-sig").splitlines()[0]]))
        else:
            book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            headers = [str(value or "").strip() for value in next(book.active.iter_rows(values_only=True))]
    except Exception as exc:
        raise HTTPException(status_code=422, detail="Unable to read the uploaded header row") from exc
    mapping = map_headers(headers)
    detected_kind = "leads" if any(field in mapping for field in ("first_name", "last_name", "full_name", "job_title")) else "accounts"
    import_kind = detected_kind if kind == "auto" else kind
    result = await create_import(workspace_id=workspace_id, import_kind=import_kind, file=file, idempotency_key=idempotency_key, user=user, db=db)
    record = db.get(Import, result["id"])
    if not record:
        raise HTTPException(status_code=500, detail="Campaign upload was not created")
    record.campaign_id = campaign.id
    if record.status in {"uploaded", "validated", "failed"}:
        job = create_job(db, workspace_id=workspace_id, created_by=user.id, job_type="campaign_prepare", name=f"Prepare {campaign.name}", idempotency_key=f"campaign-prepare:{campaign.id}:{record.id}", resource_type="campaign", resource_id=campaign.id)
        record.status = "queued"
        audit(db, workspace_id, "campaign.audience_imported", "campaign", campaign.id, {"import_id": record.id, "kind": import_kind, "detected_kind": detected_kind, "job_id": job.id}, user.id)
        db.commit(); process_import.delay(record.id)
        return {"import": import_summary(record), "job": job_data(job), "detected_kind": detected_kind}
    db.commit(); return {"import": import_summary(record), "detected_kind": detected_kind}


@app.post("/api/v1/campaigns/{campaign_id}/prepare", dependencies=[Depends(csrf_protected)])
def retry_campaign_preparation(campaign_id: str, payload: WorkspaceOnlyInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    campaign = db.get(Campaign, campaign_id)
    if not campaign or campaign.workspace_id != payload.workspace_id:
        raise HTTPException(status_code=404, detail="Campaign not found")
    records = db.query(Import).filter(Import.campaign_id == campaign.id).all()
    # Earlier campaign uploads may have completed before the campaign-audience
    # association existed. Rebuild the association idempotently from their
    # durable import rows before retrying only the incomplete work.
    from .tasks import attach_import_audience
    for record in records:
        attach_import_audience(db, record)
        if record.status in {"validated", "failed"}:
            record.status = "queued"; process_import.delay(record.id)
    audit(db, payload.workspace_id, "campaign.preparation_retried", "campaign", campaign.id, {"imports": [item.id for item in records]}, user.id)
    db.commit(); return campaign_workspace_data(db, campaign)


def replace_campaign_steps(db: Session, campaign: Campaign, steps: list[CampaignStepInput]) -> None:
    positions = [step.position for step in steps]
    if sorted(positions) != list(range(1, len(positions) + 1)):
        raise HTTPException(status_code=422, detail="Sequence steps must use consecutive positions starting at 1")
    db.query(CampaignStep).filter(CampaignStep.campaign_id == campaign.id).delete()
    for step in steps:
        db.add(CampaignStep(campaign_id=campaign.id, position=step.position, delay_hours=step.delay_hours, subject=step.subject, body=step.body, facts_used=step.facts_used))


@app.put("/api/v1/campaigns/{campaign_id}/sequence", dependencies=[Depends(csrf_protected)])
def update_campaign_sequence(campaign_id: str, payload: CampaignSequenceUpdate, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    campaign = db.get(Campaign, campaign_id)
    if not campaign or campaign.workspace_id != payload.workspace_id:
        raise HTTPException(status_code=404, detail="Campaign not found")
    replace_campaign_steps(db, campaign, payload.steps)
    campaign.audience_filter = {**(campaign.audience_filter or {}), "campaign_brief": payload.model_dump(exclude={"workspace_id", "steps"})}
    audit(db, payload.workspace_id, "campaign.sequence_updated", "campaign", campaign.id, {"preset": payload.preset, "step_count": len(payload.steps)}, user.id)
    db.commit(); return campaign_workspace_data(db, campaign)


@app.put("/api/v1/campaigns/{campaign_id}/delivery", dependencies=[Depends(csrf_protected)])
def update_campaign_delivery(campaign_id: str, payload: CampaignDeliveryUpdate, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    campaign = db.get(Campaign, campaign_id)
    if not campaign or campaign.workspace_id != payload.workspace_id:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if payload.mailbox_id:
        mailbox = db.get(MailboxConnection, payload.mailbox_id)
        if not mailbox or mailbox.workspace_id != payload.workspace_id or mailbox.status != "connected":
            raise HTTPException(status_code=422, detail="Choose a connected mailbox in this workspace")
    campaign.mailbox_id = payload.mailbox_id; campaign.timezone = payload.timezone; campaign.business_hours = payload.business_hours
    campaign.daily_limit = payload.daily_limit; campaign.per_domain_limit = payload.per_domain_limit
    audit(db, payload.workspace_id, "campaign.delivery_updated", "campaign", campaign.id, {"mailbox_id": payload.mailbox_id, "daily_limit": payload.daily_limit}, user.id)
    db.commit(); return campaign_workspace_data(db, campaign)


@app.post("/api/v1/campaigns/{campaign_id}/sequence/generate", dependencies=[Depends(csrf_protected)])
def generate_campaign_sequence(campaign_id: str, payload: CampaignSequenceGenerateInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    campaign = db.get(Campaign, campaign_id)
    if not campaign or campaign.workspace_id != payload.workspace_id:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if not settings.ollama_model:
        raise HTTPException(status_code=503, detail="OLLAMA_MODEL is required to draft a sequence")
    context = payload.model_dump(exclude={"workspace_id"})
    prompt = """You write concise B2B cold-email sequences for SignalFlow. Return only JSON matching the schema. Use only the provided campaign brief. Do not invent customers, outcomes, relationships, metrics, people, or company facts. The sequence must contain 2 or 3 email steps, include a direct CTA, and use the requested proof only if present. Facts used must name evidence from the campaign brief.\nCAMPAIGN BRIEF:\n""" + json.dumps(context)
    try:
        response = httpx.post(f"{settings.ollama_base_url.rstrip('/')}/api/generate", json={"model": settings.ollama_model, "prompt": prompt, "stream": False, "format": CampaignSequenceOutput.model_json_schema(), "options": {"temperature": 0.2}}, timeout=settings.ollama_timeout_seconds)
        response.raise_for_status()
        output = CampaignSequenceOutput.model_validate_json(response.json().get("response", "{}"))
    except Exception as exc:
        raise HTTPException(status_code=503, detail="Local Ollama could not generate a valid campaign sequence") from exc
    steps = [step.model_copy(update={"position": index, "delay_hours": 0 if index == 1 else 72 if index == 2 else 168}) for index, step in enumerate(output.steps, start=1)]
    replace_campaign_steps(db, campaign, steps)
    campaign.audience_filter = {**(campaign.audience_filter or {}), "campaign_brief": context}
    audit(db, payload.workspace_id, "campaign.sequence_generated", "campaign", campaign.id, {"model": settings.ollama_model, "preset": payload.preset}, user.id)
    db.commit(); return campaign_workspace_data(db, campaign)


@app.post("/api/v1/campaigns/{campaign_id}/audience/{member_id}/{action}", dependencies=[Depends(csrf_protected)])
def update_campaign_audience_member(campaign_id: str, member_id: str, action: str, payload: WorkspaceOnlyInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    if action not in {"include", "exclude"}:
        raise HTTPException(status_code=404, detail="Unsupported audience action")
    member = db.get(CampaignAudienceMember, member_id)
    campaign = db.get(Campaign, campaign_id)
    if not campaign or not member or campaign.workspace_id != payload.workspace_id or member.campaign_id != campaign.id:
        raise HTTPException(status_code=404, detail="Campaign audience member not found")
    member.selected = action == "include"; member.exclusion_reason = None if member.selected else "manual_exclusion"
    audit(db, payload.workspace_id, f"campaign.audience_{action}", "campaign", campaign.id, {"member_id": member.id}, user.id)
    db.commit(); return campaign_workspace_data(db, campaign)


@app.post("/api/v1/campaigns/{campaign_id}/approve", dependencies=[Depends(csrf_protected)])
def approve_campaign(campaign_id: str, payload: CampaignActionInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager")
    campaign = db.get(Campaign, campaign_id)
    if not campaign or campaign.workspace_id != payload.workspace_id:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.status not in {"draft", "approved"}:
        raise HTTPException(status_code=409, detail="Only draft campaigns can be approved")
    campaign.status = "approved"; campaign.approved_by = user.id; campaign.approved_at = utcnow()
    audit(db, payload.workspace_id, "campaign.approved", "campaign", campaign.id, {"delivery": "requires test and activation"}, user.id)
    db.commit(); return campaign_data(db, campaign)


@app.post("/api/v1/campaigns/{campaign_id}/activate", dependencies=[Depends(csrf_protected)])
def activate_campaign(campaign_id: str, payload: CampaignActionInput, idempotency_key: str = Header(..., alias="Idempotency-Key"), user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager")
    campaign = db.get(Campaign, campaign_id)
    if not campaign or campaign.workspace_id != payload.workspace_id:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.status not in {"approved", "scheduled", "paused"}:
        raise HTTPException(status_code=409, detail="Approve this campaign before activation")
    if not campaign.mailbox_id:
        raise HTTPException(status_code=409, detail="Connect and select a mailbox before activation")
    mailbox = db.get(MailboxConnection, campaign.mailbox_id)
    if not mailbox or mailbox.status != "connected":
        raise HTTPException(status_code=409, detail="Selected mailbox needs to be connected")
    if not campaign.test_sent_at:
        raise HTTPException(status_code=409, detail="Send a test email before activation")
    if not audience_leads(db, campaign):
        raise HTTPException(status_code=409, detail="No campaign-ready recipients. Qualify contacts, complete Ollama analysis, and approve their outreach drafts first.")
    job = create_job(db, workspace_id=payload.workspace_id, created_by=user.id, job_type="campaign_enrollment", name=f"Enroll {campaign.name}", idempotency_key=f"campaign-activate:{campaign.id}:{idempotency_key}", resource_type="campaign", resource_id=campaign.id)
    if job.status == "queued":
        campaign.status = "running"; campaign.activated_at = utcnow()
        db.add(OutboxEvent(topic="campaign.enrollment_requested", payload={"campaign_id": campaign.id, "workspace_id": payload.workspace_id, "job_id": job.id}))
        audit(db, payload.workspace_id, "campaign.activated", "campaign", campaign.id, {"job_id": job.id}, user.id)
    db.commit(); return {"campaign": campaign_data(db, campaign), "job": job_data(job)}


def set_campaign_state(campaign_id: str, action: str, payload: CampaignActionInput, user: User, db: Session):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager")
    campaign = db.get(Campaign, campaign_id)
    if not campaign or campaign.workspace_id != payload.workspace_id:
        raise HTTPException(status_code=404, detail="Campaign not found")
    desired = {"pause": "paused", "resume": "running", "cancel": "cancelled"}[action]
    campaign.status = desired
    if action == "cancel":
        db.query(ScheduledMessage).filter(ScheduledMessage.campaign_id == campaign.id, ScheduledMessage.status == "scheduled").update({"status": "cancelled"})
    audit_action = {"pause": "campaign.paused", "resume": "campaign.resumed", "cancel": "campaign.cancelled"}[action]
    audit(db, payload.workspace_id, audit_action, "campaign", campaign.id, {}, user.id)
    db.commit(); return campaign_data(db, campaign)


@app.post("/api/v1/campaigns/{campaign_id}/pause", dependencies=[Depends(csrf_protected)])
def pause_campaign(campaign_id: str, payload: CampaignActionInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    return set_campaign_state(campaign_id, "pause", payload, user, db)


@app.post("/api/v1/campaigns/{campaign_id}/resume", dependencies=[Depends(csrf_protected)])
def resume_campaign(campaign_id: str, payload: CampaignActionInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    return set_campaign_state(campaign_id, "resume", payload, user, db)


@app.post("/api/v1/campaigns/{campaign_id}/cancel", dependencies=[Depends(csrf_protected)])
def cancel_campaign(campaign_id: str, payload: CampaignActionInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    return set_campaign_state(campaign_id, "cancel", payload, user, db)


@app.post("/api/v1/campaigns/{campaign_id}/test", dependencies=[Depends(csrf_protected)])
def test_campaign(campaign_id: str, payload: CampaignActionInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    campaign = db.get(Campaign, campaign_id)
    if not campaign or campaign.workspace_id != payload.workspace_id:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if not campaign.mailbox_id:
        raise HTTPException(status_code=409, detail="Select a connected mailbox before sending a test")
    if campaign.status != "approved":
        raise HTTPException(status_code=409, detail="Approve the campaign before sending its test email")
    mailbox = db.get(MailboxConnection, campaign.mailbox_id)
    sample_leads = audience_leads(db, campaign)
    sample_lead = sample_leads[0] if sample_leads else None
    sample_draft = db.query(OutreachDraft).filter(OutreachDraft.lead_id == sample_lead.id, OutreachDraft.status == "approved").order_by(OutreachDraft.version.desc()).first() if sample_lead else None
    first_step = next((item for item in (sample_draft.sequence if sample_draft else []) if int(item.get("step", 0)) == 1), None)
    if not first_step:
        step = db.query(CampaignStep).filter(CampaignStep.campaign_id == campaign.id, CampaignStep.position == 1).first()
        first_step = {"subject": step.subject, "body": step.body, "facts_used": step.facts_used or []} if step else None
    if not mailbox or mailbox.status != "connected" or not first_step:
        raise HTTPException(status_code=409, detail="Connect a mailbox and approve at least one Ollama-generated recipient sequence before testing")
    recipient = payload.test_recipient or mailbox.email
    try:
        subject = f"[TEST] {first_step['subject']}"; body = first_step["body"]
        if first_step.get("cta") and first_step["cta"] not in body:
            body = f"{body.rstrip()}\n\n{first_step['cta']}"
        if mailbox.provider == "smtp_imap":
            result = smtp_imap_send(settings, to_email=recipient, subject=subject, body=body)
        else:
            from .gmail import access_token, decrypt_refresh_token, send_message
            token = access_token(settings, decrypt_refresh_token(mailbox.encrypted_refresh_token, settings))
            result = send_message(token, to_email=recipient, subject=subject, body=body)
        db.add(Message(workspace_id=payload.workspace_id, mailbox_id=mailbox.id, campaign_id=campaign.id, lead_id=sample_lead.id, direction="outbound", gmail_message_id=result.get("id"), gmail_thread_id=result.get("threadId"), to_email=recipient, subject=subject, body=body, delivery_status="test_sent"))
        campaign.test_sent_at = utcnow(); audit(db, payload.workspace_id, "campaign.test_sent", "campaign", campaign.id, {"recipient": recipient, "sample_lead_id": sample_lead.id, "ai_draft_id": sample_draft.id if sample_draft else None, "facts_used": first_step.get("facts_used", [])}, user.id); db.commit()
        return {"status": "sent", "recipient": recipient}
    except Exception as exc:
        raise HTTPException(status_code=502, detail="Mailbox test delivery failed. Check the mailbox connection and configuration.") from exc


@app.get("/api/v1/suppressions")
def list_suppressions(workspace_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    return [{"id": item.id, "email": item.normalized_email, "reason": item.reason, "created_at": item.created_at} for item in db.query(Suppression).filter(Suppression.workspace_id == workspace_id).order_by(Suppression.created_at.desc()).limit(200).all()]


@app.get("/api/v1/outreach-inbox")
def outreach_inbox(workspace_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    rows = db.query(Message).filter(Message.workspace_id == workspace_id, Message.direction == "inbound").order_by(Message.created_at.desc()).limit(200).all()
    return [{"id": item.id, "campaign_id": item.campaign_id, "lead_id": item.lead_id, "gmail_thread_id": item.gmail_thread_id, "subject": item.subject, "classification": item.reply_classification or "unknown", "received_at": item.created_at} for item in rows]


@app.get("/api/v1/analytics/overview")
def analytics_overview(workspace_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    leads = db.query(Lead).filter(Lead.workspace_id == workspace_id)
    messages = db.query(Message).filter(Message.workspace_id == workspace_id)
    inbound = messages.filter(Message.direction == "inbound")
    return {
        "accounts": db.query(Company).filter(Company.workspace_id == workspace_id).count(),
        "leads": leads.count(),
        "qualified": leads.filter(Lead.qualification == "qualified").count(),
        "campaign_ready": db.query(OutreachDraft).join(Lead, OutreachDraft.lead_id == Lead.id).filter(Lead.workspace_id == workspace_id, OutreachDraft.status == "approved").count(),
        "campaigns": db.query(Campaign).filter(Campaign.workspace_id == workspace_id).count(),
        "enrolled": db.query(CampaignEnrollment).join(Campaign, CampaignEnrollment.campaign_id == Campaign.id).filter(Campaign.workspace_id == workspace_id).count(),
        "sent": messages.filter(Message.direction == "outbound", Message.delivery_status == "sent").count(),
        "replied": inbound.count(),
        "interested": inbound.filter(Message.reply_classification == "interested").count(),
        "bounced": inbound.filter(Message.reply_classification == "hard_bounce").count(),
        "unsubscribed": inbound.filter(Message.reply_classification == "unsubscribe").count(),
        "delivery_note": "Gmail confirms API acceptance, not final delivery. Open tracking is not implemented.",
    }


@app.post("/api/v1/suppressions", dependencies=[Depends(csrf_protected)])
def create_suppression(payload: SuppressionInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    from .services import normalize_email
    email, error = normalize_email(payload.email)
    if error or not email:
        raise HTTPException(status_code=422, detail="Provide a valid email address")
    record = db.query(Suppression).filter(Suppression.workspace_id == payload.workspace_id, Suppression.normalized_email == email).first()
    if not record:
        record = Suppression(workspace_id=payload.workspace_id, normalized_email=email, reason=payload.reason, created_by=user.id); db.add(record)
        db.query(Lead).filter(Lead.workspace_id == payload.workspace_id, Lead.normalized_email == email).update({"qualification": "suppressed", "next_action": "do_not_contact"})
        audit(db, payload.workspace_id, "suppression.created", "suppression", email, {"reason": payload.reason}, user.id)
    db.commit(); return {"id": record.id, "email": record.normalized_email, "reason": record.reason}


@app.post("/api/v1/exports", dependencies=[Depends(csrf_protected)])
def create_export(payload: ExportInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep", "viewer")
    if payload.kind == "processed":
        rows = [{"id": lead.id, "name": lead.full_name, "email": lead.email, "score": lead.score, "qualification": lead.qualification, "next_action": lead.next_action, "owner_id": lead.owner_id} for lead in db.query(Lead).filter(Lead.workspace_id == payload.workspace_id).all()]
    elif payload.kind == "rejected":
        rows = [{"import_id": row.import_id, "row_number": row.row_number, "outcome": row.outcome, "errors": json.dumps(row.errors)} for row in db.query(ImportRow).join(Import).filter(Import.workspace_id == payload.workspace_id, ImportRow.outcome.in_(["rejected", "needs_review"])).all()]
    elif payload.kind == "audit":
        rows = [{"action": row.action, "resource_type": row.resource_type, "resource_id": row.resource_id, "payload": json.dumps(row.payload), "created_at": row.created_at.isoformat()} for row in db.query(AuditLog).filter(AuditLog.workspace_id == payload.workspace_id).all()]
    else:
        rows = [{"lead_id": draft.lead_id, "email": lead.email if lead else None, "name": lead.full_name if lead else None, "subject": draft.subject, "body": draft.body, "sequence": json.dumps(draft.sequence), "status": draft.status, "version": draft.version, "created_at": draft.created_at.isoformat()} for draft, lead in db.query(OutreachDraft, Lead).outerjoin(Lead, Lead.id == OutreachDraft.lead_id).filter(OutreachDraft.workspace_id == payload.workspace_id).all()]
    headers = list(rows[0].keys()) if rows else ["message"]
    key = f"exports/{payload.workspace_id}/{payload.kind}-{secrets.token_hex(8)}.csv"
    put_bytes(key, safe_csv(rows or [{"message": "No rows"}], headers), "text/csv")
    audit(db, payload.workspace_id, "export.created", "export", key, {"kind": payload.kind, "row_count": len(rows)}, user.id); db.commit()
    return {"kind": payload.kind, "download_url": download_url(key), "row_count": len(rows)}


@app.get("/api/v1/audit-logs")
def audit_logs(workspace_id: str, limit: int = 100, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    return [{"action": item.action, "resource_type": item.resource_type, "resource_id": item.resource_id, "payload": item.payload, "created_at": item.created_at} for item in db.query(AuditLog).filter(AuditLog.workspace_id == workspace_id).order_by(AuditLog.created_at.desc()).limit(min(limit, 200)).all()]


@app.get("/api/v1/integrations/ollama/test")
def ollama_status(user: User = Depends(current_user)):
    if not settings.ollama_model:
        return {"required": True, "ready": False, "configured_model": None, "detail": "OLLAMA_MODEL is required"}
    try:
        response = httpx.get(f"{settings.ollama_base_url.rstrip('/')}/api/tags", timeout=min(settings.ollama_timeout_seconds, 5))
        response.raise_for_status()
        models = {item.get("name") for item in response.json().get("models", [])}
        ready = settings.ollama_model in models
        return {"required": True, "ready": ready, "configured_model": settings.ollama_model, "detail": "Ready" if ready else "Configured model is not installed in Ollama"}
    except httpx.HTTPError:
        return {"required": True, "ready": False, "configured_model": settings.ollama_model, "detail": "Ollama is unavailable at the configured local endpoint"}


@app.get("/api/v1/integrations")
def integrations(workspace_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    """List supported adapters and workspace-local configuration state.

    Credential material is intentionally excluded from every response.
    """
    workspace_membership(workspace_id, user, db)
    records = {item.provider: item for item in db.query(IntegrationConnection).filter(IntegrationConnection.workspace_id == workspace_id).all()}
    return {"integrations": [public_connection(records.get(item.key), item) for item in registry.definitions()]}


@app.put("/api/v1/integrations/{provider}", dependencies=[Depends(csrf_protected)])
def configure_integration(provider: str, payload: IntegrationConfigurationInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    adapter = registry.get(provider)
    if not adapter:
        raise HTTPException(status_code=404, detail="Unsupported integration provider")
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager")
    connection = integration_connection_for(db, payload.workspace_id, adapter.definition.key)
    if not connection:
        connection = IntegrationConnection(workspace_id=payload.workspace_id, provider=adapter.definition.key, created_by=user.id)
        db.add(connection); db.flush()
    connection.configuration = payload.configuration
    connection.capabilities = list(adapter.definition.capabilities)
    if payload.credentials is not None:
        if not payload.credentials or any(not key.strip() or not value for key, value in payload.credentials.items()):
            raise HTTPException(status_code=422, detail="Credentials must contain non-empty keys and values")
        try:
            connection.encrypted_credentials = encrypt_refresh_token(json.dumps(payload.credentials, sort_keys=True), settings)
        except GmailConfigurationError as exc:
            raise HTTPException(status_code=503, detail=str(exc)) from exc
    connection.status = "configured" if connection.encrypted_credentials else "needs_credentials"
    connection.last_error = None
    audit(db, payload.workspace_id, "integration.configured", "integration", connection.id, {"provider": adapter.definition.key, "credential_fields_received": sorted(payload.credentials or {})}, user.id)
    db.commit(); db.refresh(connection)
    return public_connection(connection, adapter.definition)


@app.post("/api/v1/integrations/{provider}/test", dependencies=[Depends(csrf_protected)])
def test_integration(provider: str, payload: IntegrationTestInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    adapter = registry.get(provider)
    if not adapter:
        raise HTTPException(status_code=404, detail="Unsupported integration provider")
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager")
    connection = integration_connection_for(db, payload.workspace_id, adapter.definition.key)
    result = adapter.test_connection(has_credentials=bool(connection and connection.encrypted_credentials), configuration=(connection.configuration if connection else {}))
    if connection:
        connection.status = result.status
        connection.last_test_at = utcnow()
        connection.last_error = None if result.status == "adapter_pending" else result.detail
        audit(db, payload.workspace_id, "integration.tested", "integration", connection.id, {"provider": adapter.definition.key, "status": result.status}, user.id)
        db.commit(); db.refresh(connection)
    return {"provider": adapter.definition.key, "status": result.status, "ready": result.ready, "detail": result.detail, "integration": public_connection(connection, adapter.definition)}


@app.post("/api/v1/integrations/apollo/import", status_code=status.HTTP_201_CREATED, dependencies=[Depends(csrf_protected)])
def import_apollo_people(payload: ApolloImportInput, idempotency_key: str = Header(..., alias="Idempotency-Key"), user: User = Depends(current_user), db: Session = Depends(get_db)):
    """Import a single Apollo people-search page through the canonical CSV flow."""
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    connection = integration_connection_for(db, payload.workspace_id, "apollo")
    if not connection or not connection.encrypted_credentials:
        raise HTTPException(status_code=409, detail="Configure the Apollo integration before importing prospects")
    request_fingerprint = json.dumps({"filters": payload.filters, "page": payload.page, "per_page": payload.per_page}, sort_keys=True)
    request_hash = hashlib.sha256(f"{payload.workspace_id}:apollo:{request_fingerprint}".encode()).hexdigest()
    existing = db.query(Import).filter(Import.workspace_id == payload.workspace_id, Import.idempotency_key == idempotency_key).first()
    if existing:
        if existing.request_hash != request_hash:
            raise HTTPException(status_code=409, detail="Idempotency key was used with a different Apollo search")
        jobs = resource_jobs(db, payload.workspace_id, existing.id)
        return {**import_summary(existing), "job": jobs[0] if jobs else None}
    try:
        people = apollo_people_search(connection, settings, filters=payload.filters, page=payload.page, per_page=payload.per_page)
    except ApolloConfigurationError as exc:
        connection.status = "error"; connection.last_error = str(exc); db.commit()
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    if not people:
        connection.sync_status = "idle"; connection.last_error = None; connection.last_sync_at = utcnow()
        audit(db, payload.workspace_id, "apollo.search_completed", "integration", connection.id, {"page": payload.page, "received": 0}, user.id)
        db.commit(); return {"received": 0, "detail": "Apollo returned no people for this page."}
    content = apollo_people_csv(people)
    checksum = hashlib.sha256(content).hexdigest()
    filename = f"apollo-people-page-{payload.page}.csv"
    object_key = f"imports/{payload.workspace_id}/apollo/{checksum}-{secrets.token_hex(8)}.csv"
    put_bytes(object_key, content, "text/csv")
    headers = next(__import__("csv").reader([content.decode("utf-8").splitlines()[0]]))
    record = Import(workspace_id=payload.workspace_id, created_by=user.id, filename=filename, object_key=object_key, checksum=checksum, mime_type="text/csv", size_bytes=len(content), idempotency_key=idempotency_key, request_hash=request_hash, kind="leads", status="queued", column_mapping=map_headers(headers))
    db.add(record); db.flush()
    job = create_job(db, workspace_id=payload.workspace_id, created_by=user.id, job_type="import", name=f"Apollo people import page {payload.page}", idempotency_key=f"apollo-import:{record.id}", resource_type="import", resource_id=record.id, counters={"total": len(people), "processed": 0})
    connection.status = "configured"; connection.sync_status = "queued"; connection.last_error = None; connection.last_sync_at = utcnow()
    audit(db, payload.workspace_id, "apollo.import_queued", "import", record.id, {"integration_id": connection.id, "page": payload.page, "received": len(people), "job_id": job.id}, user.id)
    db.commit()
    process_import.delay(record.id)
    return {**import_summary(record), "received": len(people), "job": job_data(job)}


@app.get("/api/v1/integrations/gmail")
def gmail_connections(workspace_id: str, user: User = Depends(current_user), db: Session = Depends(get_db)):
    workspace_membership(workspace_id, user, db)
    configured = bool(settings.gmail_client_id and settings.gmail_client_secret and settings.credential_encryption_key)
    return {"configured": configured, "redirect_uri": settings.gmail_redirect_uri, "connections": [mailbox_data(item) for item in db.query(MailboxConnection).filter(MailboxConnection.workspace_id == workspace_id).order_by(MailboxConnection.created_at.desc()).all()], "setup_required": [] if configured else ["GMAIL_CLIENT_ID", "GMAIL_CLIENT_SECRET", "CREDENTIAL_ENCRYPTION_KEY"], "smtp_imap_configured": smtp_imap_configured(settings), "smtp_imap_setup_required": smtp_imap_setup_required(settings)}


@app.post("/api/v1/integrations/smtp-imap/connect", dependencies=[Depends(csrf_protected)])
def smtp_imap_connect(payload: WorkspaceOnlyInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager")
    try:
        email = verify_smtp_imap(settings)
    except SmtpImapConfigurationError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    mailbox = db.query(MailboxConnection).filter(MailboxConnection.workspace_id == payload.workspace_id, MailboxConnection.email == email).first()
    encrypted_password = encrypt_refresh_token(settings.smtp_password, settings)
    if mailbox:
        mailbox.provider = "smtp_imap"; mailbox.encrypted_refresh_token = encrypted_password; mailbox.scopes = ["smtp.send", "imap.read"]; mailbox.status = "connected"; mailbox.last_error = None
    else:
        mailbox = MailboxConnection(workspace_id=payload.workspace_id, connected_by=user.id, provider="smtp_imap", email=email, encrypted_refresh_token=encrypted_password, scopes=["smtp.send", "imap.read"])
        db.add(mailbox); db.flush()
    audit(db, payload.workspace_id, "smtp_imap.connected", "mailbox", mailbox.id, {"email": email, "provider": "smtp_imap"}, user.id)
    db.commit(); return {"mailbox": mailbox_data(mailbox)}


@app.post("/api/v1/integrations/gmail/connect", dependencies=[Depends(csrf_protected)])
def gmail_connect(payload: WorkspaceOnlyInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager")
    state = secrets.token_urlsafe(32); verifier = code_verifier()
    try:
        url = oauth_url(settings, state, verifier)
    except GmailConfigurationError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    db.add(GmailOAuthState(state=state, workspace_id=payload.workspace_id, user_id=user.id, code_verifier=verifier, expires_at=utcnow() + timedelta(minutes=10)))
    audit(db, payload.workspace_id, "gmail.oauth_started", "integration", "gmail", {}, user.id)
    db.commit(); return {"authorization_url": url, "expires_in_seconds": 600}


@app.get("/api/v1/integrations/gmail/callback")
def gmail_callback(state: str, code: str | None = None, error: str | None = None, db: Session = Depends(get_db)):
    record = db.query(GmailOAuthState).filter(GmailOAuthState.state == state).first()
    target = f"{settings.public_app_url.rstrip('/')}/app/integrations"
    if not record or record.expires_at <= utcnow():
        return RedirectResponse(f"{target}?gmail=expired")
    db.delete(record)
    if error or not code:
        db.commit(); return RedirectResponse(f"{target}?gmail=cancelled")
    try:
        tokens = exchange_code(settings, code, record.code_verifier)
        email = google_email(tokens["access_token"])
        connection = db.query(MailboxConnection).filter(MailboxConnection.workspace_id == record.workspace_id, MailboxConnection.email == email).first()
        if connection:
            connection.encrypted_refresh_token = encrypt_refresh_token(tokens["refresh_token"], settings); connection.scopes = tokens.get("scope", "").split(); connection.status = "connected"; connection.last_error = None
        else:
            connection = MailboxConnection(workspace_id=record.workspace_id, connected_by=record.user_id, email=email, encrypted_refresh_token=encrypt_refresh_token(tokens["refresh_token"], settings), scopes=tokens.get("scope", "").split())
            db.add(connection); db.flush()
        audit(db, record.workspace_id, "gmail.connected", "mailbox", connection.id, {"email": email, "scopes": connection.scopes}, record.user_id)
        db.commit(); return RedirectResponse(f"{target}?gmail=connected")
    except Exception:
        db.rollback(); return RedirectResponse(f"{target}?gmail=failed")


@app.post("/api/v1/integrations/gmail/{mailbox_id}/sync", dependencies=[Depends(csrf_protected)])
def gmail_sync(mailbox_id: str, payload: WorkspaceOnlyInput, idempotency_key: str = Header(..., alias="Idempotency-Key"), user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager", "rep")
    mailbox = db.get(MailboxConnection, mailbox_id)
    if not mailbox or mailbox.workspace_id != payload.workspace_id:
        raise HTTPException(status_code=404, detail="Mailbox not found")
    job = create_job(db, workspace_id=payload.workspace_id, created_by=user.id, job_type="gmail_sync", name=f"Sync {mailbox.email}", idempotency_key=f"gmail-sync:{mailbox.id}:{idempotency_key}", resource_type="mailbox", resource_id=mailbox.id)
    if job.status == "queued":
        db.add(OutboxEvent(topic="gmail.sync_requested", payload={"mailbox_id": mailbox.id, "workspace_id": payload.workspace_id, "job_id": job.id}))
        audit(db, payload.workspace_id, "gmail.sync_queued", "mailbox", mailbox.id, {"job_id": job.id}, user.id)
    db.commit(); return {"job": job_data(job)}


@app.post("/api/v1/integrations/gmail/{mailbox_id}/disconnect", dependencies=[Depends(csrf_protected)])
def gmail_disconnect(mailbox_id: str, payload: WorkspaceOnlyInput, user: User = Depends(current_user), db: Session = Depends(get_db)):
    membership = workspace_membership(payload.workspace_id, user, db); require_roles(membership, "admin", "manager")
    mailbox = db.get(MailboxConnection, mailbox_id)
    if not mailbox or mailbox.workspace_id != payload.workspace_id:
        raise HTTPException(status_code=404, detail="Mailbox not found")
    mailbox.status = "disconnected"; mailbox.encrypted_refresh_token = ""
    audit(db, payload.workspace_id, "gmail.disconnected", "mailbox", mailbox.id, {}, user.id)
    db.commit(); return {"id": mailbox.id, "status": mailbox.status}
